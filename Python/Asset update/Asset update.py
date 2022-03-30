#import xlrd
import csv
import pandas as pd

old_data = "Old_Data.csv"
new_data = "New_Data.csv"

changes = open("Update_Assets.csv", newline="")
reader = csv.reader(changes)
header = next(reader)
data = [row for row in reader]

rows_changes = []

for i in  range(len(data)):
    row = {header[0]:data[i][0], # PC NAME
           header[1]:data[i][1]} # User
    rows_changes.append(row)

#pc_name = input("Enter PC Name: ")#"GBRLW-00DHWZLQ2"
#user_name = input("Enter username: ")#"Hoopla!"

df = pd.DataFrame(pd.read_excel("Dell Assets.xlsx"))
df.to_csv("Old_Data.csv", index=False)# copy old .xlsx to a .csv

file = open(old_data, newline='')
reader = csv.reader(file)
header = next(reader) #first line is the header
data = [row for row in reader] # 2D list containing all values

rows = [] # 1D list for dictionaries containing headers and values

for i in range(len(data)): # necessary for writer.writerows to use the dictionary key/value pairs as tuples
    row = {header[0]:data[i][0], # PC NAME
           header[1]:data[i][1], # Model
           header[2]:data[i][2], # Date assigned
           header[3]:data[i][3], # User
           header[4]:data[i][4], # Docking Station
           header[5]:data[i][5][:-9], # End date 
           header[6]:data[i][6], # Comment
           header[7]:data[i][7]} # Company 
    rows.append(row)

for i in range(len(rows)):
    for idx, (key, value) in enumerate(rows[i].items()):
        if rows_changes[idx][header[0]] in rows[i].values():
            rows[i] = {header[0]:rows_changes[idx][header[0]], # PC NAME
                        header[1]:data[i][1], # Model
                        header[2]:data[i][2], # Date assigned
                        header[3]:rows_changes[idx][header[3]], # User
                        header[4]:data[i][4], # Docking Station
                        header[5]:data[i][5][:-9], # End date 
                        header[6]:data[i][6], # Comment
                        header[7]:data[i][7]} # Company 

with open(new_data, 'w', encoding='UTF8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=header)
    writer.writeheader()
    writer.writerows(rows)

#------------------------Pandas to write to .xlsx---------------------------------------------------------------

csvFile = pd.read_csv(new_data)

with pd.ExcelWriter("Excel_Test.xlsx") as excelWriter:
    csvFile.to_excel(
    excelWriter,
    index = False
    )

#------------------------OpenPyXl to write to .xlsx--------------------------------------------------------------

import openpyxl as px
# from copy import copy#, deepcopy
# from openpyxl.styles import *


# old_wb = px.load_workbook("DELL ASSETS.xlsx")
new_wb = px.load_workbook("Excel_Test.xlsx")

# newStyle = NamedStyle(name="AssetList")

# newStyle.font = copy(old_wb.font)

# old_sheet = old_wb.active
new_sheet = new_wb.active

# for styl in 

# new_sheet = copy(old_sheet)


# for row in default_sheet.rows:
#     col_idx = float(default_sheet.get_highest_column())
# starting_col = chr(65 + int(col_idx))
# for row in old_sheet.rows:
#     for cell in row:
#         new_sheet[cell.get_coordinate()] = cell.value
#         <copy also style of each cell>
#         old_wb._styles[cell.get_coordinate()] = copy(
#         old_sheet._styles[cell.get_coordinate()])



new_sheet.auto_filter.ref = f"A1:H{str(len(data)+1)}"

new_wb.save("Excel_Test.xlsx")

# wb = Workbook()
# sheet = wb.active
# sheet.title = "data"

# wb.save("testing.xlsx")

# sheet["A1"] = header[0] # PC NAME
# sheet["B1"] = header[1] # Model
# sheet["C1"] = header[2] # Date assigned
# sheet["D1"] = header[3] # User
# sheet["E1"] = header[4] # Docking Station
# sheet["F1"] = header[5] # End date 
# sheet["G1"] = header[6] # Comment
# sheet["H1"] = header[7] # Company 

# for idx,pc in enumerate(data, start=1):
#     c = sheet.cell(row = idx, column = 1)
#     c.value = pc[header[0]]
#     i = str(idx)
#     sheet["A" + i] = header[0] # PC NAME
#     sheet["B" + i] = header[1] # Model
#     sheet["C" + i] = header[2] # Date assigned
#     sheet["D" + i] = header[3] # User
#     sheet["E" + i] = header[4] # Docking Station
#     sheet["F" + i] = header[5] # End date 
#     sheet["G" + i] = header[6] # Comment
#     sheet["H" + i] = header[7] # Company 
